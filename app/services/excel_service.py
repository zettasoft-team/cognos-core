"""
시트 데이터 ↔ Excel (.xlsx) 변환
"""
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

SHEET_LABELS = {
    "table_list":     "테이블 목록",
    "column_mapping": "컬럼 매핑 상세",
    "join_relation":  "조인 관계",
    "dimension_view": "DimensionalView",
    "meaning_dict":   "NL2SQL 의미사전",
    "fact_calc":      "Fact 산출식",
}

HEADER_FILL  = PatternFill("solid", fgColor="1A1917")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def sheets_to_excel(all_sheets: dict[str, dict]) -> bytes:
    """시트 데이터 dict → Excel bytes"""
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    for sheet_id, label in SHEET_LABELS.items():
        sheet_data = all_sheets.get(sheet_id, {"columns": [], "rows": []})
        ws = wb.create_sheet(title=label)

        # 헤더
        headers = sheet_data.get("columns", [])
        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill  = HEADER_FILL
            cell.font  = HEADER_FONT
            cell.alignment = HEADER_ALIGN

        # 데이터
        for ri, row in enumerate(sheet_data.get("rows", []), start=2):
            for ci, val in enumerate(row, start=1):
                ws.cell(row=ri, column=ci, value=val)

        # 컬럼 너비 자동 조정
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def excel_to_sheets(file_bytes: bytes) -> dict[str, dict]:
    """Excel bytes → 시트 데이터 dict"""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)

    # 시트명(한글) → sheet_id 역매핑
    label_to_id = {v: k for k, v in SHEET_LABELS.items()}
    result = {}

    for ws in wb.worksheets:
        sheet_id = label_to_id.get(ws.title)
        if not sheet_id:
            continue

        rows_iter = list(ws.iter_rows(values_only=True))
        if not rows_iter:
            continue

        columns = [str(c) for c in rows_iter[0]]
        rows    = [[str(c) if c is not None else "" for c in row]
                   for row in rows_iter[1:]]

        result[sheet_id] = {"columns": columns, "rows": rows}

    return result
